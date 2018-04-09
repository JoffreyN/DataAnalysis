import os,sys,time,re
from openpyxl import Workbook
path=os.getcwd()+'\\'
filename=input('输入文件名(脚本所在文件夹)：')

'''
V1.3
修复bug：当请求长度小于100时数据被丢弃(getallList()中的判断语句增加正则匹配)
--------------------------------------------------
V1.2.1
优化getallDic中进度条
--------------------------------------------------
V1.2
优化进度条
优化函数getallList()逻辑：优化请求后面有postdata时的逻辑
优化函数onestrTodic()逻辑：优化有postdata时的逻辑
优化函数saveExcel()逻辑：写自定义字段时的逻辑
--------------------------------------------------
V1.1
修复BUG:函数onestrTodic()逻辑不严谨，导致数据写入错行
--------------------------------------------------
V1.0
把库xlwt换为openpyxl库
--------------------------------------------------
v0.9
初步编写完成
'''
def jindutiao(i,count,word='处理数据'):
    words='正在{}：'.format(word)
    width=(os.get_terminal_size().columns-len(words))//2
    m='-'*(((i+1)*width)//count)  
    sys.stdout.write(words+'|'+m+'>'+' '*(width-((i+1)*width)//count)+'|'+str(round((i+1)/count*100,2))+"%\r")
    sys.stdout.flush()
    #time.sleep(0.01)

def getallLines(path):
    print('\n正在读取文件...')
    with open(path,'r',encoding='utf-8',errors='ignore') as f:
        allLines=f.readlines()
    return allLines

def getallList(allLines):
    strLines=''.join(allLines)
    allList=strLines.split('='*54)
    count=len(allList)
    for i in range(count):
        if re.search(r'\d{1,2}:\d{1,2}:\d{1,2}  https?://',allList[i]):
            allList[i]=''
        elif len(allList[i])<=10:
            allList[i]=''
        jindutiao(i,count,'生成列表')
    print('\n正在准备进行下一步处理...')
    while '' in allList:
        allList.remove('')
    for k in range(len(allList)):
        allList[k]=allList[k].replace('\n\n','\nPostData:',1)
    return allList

def onestrTodic(onestr):
    dic={}
    deln=onestr.strip('\n')
    onelist=deln.split('\n')
    onelist[0]=onelist[0].replace(' ',':',1)
    while '' in onelist:
        onelist.remove('')
    for i in range(len(onelist)):
        if 'PostData:' in onelist[i]:
            onelist[i]='\n'.join(onelist[i:])
            dic[onelist[i].split(':',1)[0]]=onelist[i].split(':',1)[1]
            break
        else:
            dic[onelist[i].split(':',1)[0]]=onelist[i].split(':',1)[1]
    return dic

def getallDic(allList):
    allDic=[]
    x=0
    for i in allList:        
        allDic.append(onestrTodic(i))
        jindutiao(x,len(allList),'生成字典')
        x+=1
    #print(allDic)
    return allDic

def saveExcel(allDic):
    #row0是表头，如果发现某个自定义字段出现的次数较多，可将该字段添加进row0内，放在'otherKeys'之前；同理，也可以删除不想要的字段
    row0=['GET','POST','PostData','Host','Accept', 'Accept-Charset', 'Accept-Encoding',
         'Accept-Language', 'Accept-Ranges', 'Authorization', 'Cache-Control',
         'Connection', 'Cookie', 'Content-Length', 'Content-length','Content-Type',
         'Content-type','Date','Expect', 'From',  'Max-Forwards', 'Pragma',
         'Proxy-Authorization','Range', 'Referer', 'Upgrade', 'User-Agent',
         'Warning','otherKeys']
    excel=Workbook()
    sheet=excel.active
    z=0
    print('\n',end='')
    for i in range(len(allDic)):
        s=''
        for j in range(len(row0)):            
            if i==0:
                sheet.cell(row=1,column=j+1,value=row0[j])
                if row0[j] in allDic[i]:  
                    sheet.cell(row=i+2,column=j+1,value=repr(allDic[i][row0[j]].strip())[1:-1])
                else:  
                    sheet.cell(row=i+2,column=j+1,value='')
            elif row0[j] in allDic[i]:            
                sheet.cell(row=i+2,column=j+1,value=repr(allDic[i][row0[j]].strip())[1:-1])
            else:      
                sheet.cell(row=i+2,column=j+1,value='')
            jindutiao(z,len(allDic)*len(row0),'处理数据并写入文件')
            z+=1
        
        for k in list(allDic[i].keys()):
            if k not in row0:
                s=s+k+':'+allDic[i][k].strip()+';'
                sheet.cell(row=i+2,column=row0.index(row0[-1])+1,value=repr(s)[1:-1])
            else:
                continue
    print('\n正在保存...')
    excel.save(path+filename.split('.')[0]+'.xlsx')
    print('\n处理完成！文件保存至{}'.format(path+filename.split('.')[0]+'.xlsx'))

def main():       
    filepath=path+filename
    allLines=getallLines(filepath)
    allList=getallList(allLines)
    allDic=getallDic(allList)
    saveExcel(allDic)
    
if __name__=='__main__':
    main()
