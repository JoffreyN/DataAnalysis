import os,sys,time
from openpyxl import Workbook
path=os.getcwd()+'\\'
filename=input('输入文件名(脚本所在文件夹)：')

'''
V1.1
修复BUG:函数onestrTodic()逻辑不严谨，导致数据写入错行
---------------------------------------
V1.0
把库xlwt换为openpyxl库
--------------------------------------------------
v0.9
初步编写完成
'''
def jindutiao(i,count,word='处理数据'):
    m='-'*int((i+1)/(count)*50)
    sys.stdout.write('正在{}：'.format(word)+'|'+m+'>'+' '*(50-int((i+1)/(count)*50))+'|'+str(int((i+1)/(count)*100))+"%\r")
    sys.stdout.flush()
    #time.sleep(0.01)

def getallLines(path):
    print('\n正在读取文件...')
    with open(path,'r',encoding='utf-8',errors='ignore') as f:
        allLines=f.readlines()
    print('\n')
    #print(allLines)
    return allLines

def getallList(allLines):
    strLines=''.join(allLines)
    allList=strLines.split('='*54)
    count=len(allList)
    #print(count)
    for i in range(count):
        if len(allList[i])<=100:
            allList[i]=''
        jindutiao(i,count,'生成列表')
    print('\n')
    print('正在准备进行下一步处理...\n')
    while '' in allList:
        allList.remove('')
    while '\n\n\n\n' in allList:
        allList.remove('\n\n\n\n')
    while '\n\n\n' in allList:
        allList.remove('\n\n\n')
    #print(allList)
    return allList

def onestrTodic(onestr):
    dic={}
    deln=onestr.strip('\n')
    onelist=deln.split('\n')
    onelist[0]=onelist[0].replace(' ',':',1)
    while '' in onelist:
        onelist.remove('')
    #print(onelist)
    if '=' in onelist[-1]:
        onelist[-1]='PostData:'+onelist[-1]
        '''
    if not onelist[-1].startswith('Accept:'):
        if 'Accept: */*' in onelist:
            onelist[onelist.index('Accept: */*')+1]='\n'.join(onelist[onelist.index('Accept: */*')+1:])
            del onelist[onelist.index('Accept: */*')+2:]
            onelist[-1]='PostData:'+onelist[-1]
            '''
    for i in onelist:
        l=i.split(':',1)
        if len(l)==1:
            l.append(' ')      
        dic[l[0]]=l[1]
    return dic

def getallDic(allList):
    allDic=[]
    for i in allList:
        allDic.append(onestrTodic(i))
        jindutiao(allList.index(i),len(allList),'生成字典')
    print('\n')
    #print(allList)
    return allDic

def saveExcel(allDic):
    #line=len(allDic)
    row0=['GET','POST','PostData','Host','Accept', 'Accept-Charset', 'Accept-Encoding',
         'Accept-Language', 'Accept-Ranges', 'Authorization', 'Cache-Control',
         'Connection', 'Cookie', 'Content-Length', 'Content-length','Content-Type', 'Content-type','Date',
         'Expect', 'From', 'If-Match', 'If-Modified-Since', 'If-None-Match',
         'If-Range', 'If-Unmodified-Since', 'Max-Forwards', 'Pragma',
         'Proxy-Authorization', 'Range', 'Referer', 'TE', 'Upgrade',
         'User-Agent', 'Via', 'Warning','Acunetix-Aspect','Acunetix-Aspect-Password',
         'Acunetix-Aspect-Queries','otherKeys']
    excel=Workbook()
    sheet=excel.active
    z=0
    for i in range(len(allDic)):        
        for j in range(len(row0)):
            z+=1
            if i==0:
                sheet.cell(row=1,column=j+1,value=row0[j])
                if row0[j] in allDic[i]:  
                    sheet.cell(row=i+2,column=j+1,value=repr(allDic[i][row0[j]].strip())[1:-1])
                else:  
                    sheet.cell(row=i+2,column=j+1,value='')
            elif row0[j] in allDic[i]:            
                sheet.cell(row=i+2,column=j+1,value=repr(allDic[i][row0[j]].strip())[1:-1])
            else:      
                sheet.cell(row=i+2,column=j+1,value='')
            jindutiao(z,len(allDic)*len(row0),'处理数据并写入文件')
    
        for k in list(allDic[i].keys()):
            if k not in row0:               
                sheet.cell(row=i+2,column=row0.index(row0[-1])+1,value=repr(k)[1:-1]+':')             
                sheet.cell(row=i+2,column=row0.index(row0[-1])+2,value=repr(allDic[i][k].strip())[1:-1])
            else:
                continue
    print('\n')
    excel.save(path+filename.split('.')[0]+'.xlsx')
    print('处理完成！文件保存至{}'.format(path+filename.split('.')[0]+'.xlsx'))

def main():       
    filepath=path+filename
    allLines=getallLines(filepath)
    allList=getallList(allLines)
    allDic=getallDic(allList)
    saveExcel(allDic)
    #print(data)
if __name__=='__main__':
    main()
