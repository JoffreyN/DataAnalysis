import os,sys,time,re
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font
from scapy.utils import *
from scapy.route import *
from scapy.layers.all import *

def main(filepath):
    if not filepath.endswith('.pcap'):filepath=filepath+'.pcap'
    try:
        datas,excel,Num,count=PcapReader(filepath),Workbook(),0,0
        HttpStat,HttpStat.title,EpmapStat,NBSSStat,TLSStat,MySQLStat,OtherTCP,UDPStat=excel.active,'HTTP统计',excel.create_sheet("Epmap统计"),excel.create_sheet("NBSS统计"),excel.create_sheet("TLS统计"),excel.create_sheet("MySQL统计"),excel.create_sheet("OtherTCP"),excel.create_sheet("UDP统计")
        Row0(HttpFormHeader,PublicFormHeader,HttpStat,EpmapStat,NBSSStat,TLSStat,MySQLStat,OtherTCP,UDPStat)
        while True:
            packet=datas.read_packet()
            if packet is None:
                userStyles(HttpStat,EpmapStat,NBSSStat,TLSStat,MySQLStat,OtherTCP,UDPStat)
                print(filepath)
                excel.save(os.path.splitext(filepath)[0]+'.xlsx')
                print('\n处理完成！文件保存至{}'.format(os.path.splitext(filepath)[0]+'.xlsx'))
                break
            else:
                Num+=1;count+=1
                if count>12:count=1
                if packet.sprintf('%Ether.type%')=='0x800':# IP 包
                    if NullOrNot(packet)!='':
                        if packet.sprintf('%IP.proto%')=='tcp':#TCP包
                            if packet.sprintf('%TCP.dport%')in['http','8080']:HttpStat.append(ListToDic(HttpFilter(packet,Num)))#是http协议,如果http是其它端口，则需要添加
                            elif packet.sprintf('%TCP.dport%')in['epmap','135']:EpmapStat.append(PublicDic(Num,packet))#epmap服务
                            elif packet.sprintf('%TCP.dport%')in['netbios_ssn','microsoft_ds','139','445']:NBSSStat.append(PublicDic(Num,packet))#NetBIOS Session Service && SMB
                            elif packet.sprintf('%TCP.dport%')in['https']or packet.sprintf('%TCP.sport%')in['https']:TLSStat.append(PublicDic(Num,packet))#TLS
                            elif packet.sprintf('%TCP.dport%')in['3306']:MySQLStat.append(PublicDic(Num,packet))
                            else:OtherTCP.append(PublicDic(Num,packet))
                        elif packet.sprintf('%IP.proto%')=='udp':#UDP包
                            if NullOrNot(packet)!='':
                                UDPStat.append(PublicDic(Num,packet))
                ProgressBar(count,startTime)
    except Exception as e:
        print(e)

def Row0(formheader1,formheader2,*sheetname):
    for sheet in sheetname:
        if sheet.title=='HTTP统计':sheet.append(formheader1.positive)
        else:sheet.append(formheader2.positive)

def PublicDic(Num,packet):
    if packet.sprintf('%IP.proto%')=='tcp':
        if packet.sprintf('%TCP.sport%').isdigit():protocol=packet.sprintf('%TCP.dport%')
        else:protocol=packet.sprintf('%TCP.sport%')
        dic={1:Num,2:packet[TCP].sport,3:packet[TCP].dport,4:protocol,5:bOrNot(packet.sprintf('%Raw.load%')).strip('\\r\\n'),6:bOrNot(packet.sprintf('%SMBNegociate_Protocol_Request_Tail.BufferData%')),7:bOrNot(packet.sprintf('%Padding.load%'))}
    elif packet.sprintf('%IP.proto%')=='udp':
        if packet.sprintf('%UDP.sport%').isdigit():protocol=packet.sprintf('%UDP.dport%')
        else:protocol=packet.sprintf('%UDP.sport%')
        dic={1:Num,2:packet[UDP].sport,3:packet[UDP].dport,4:protocol,5:bOrNot(packet.sprintf('%Raw.load%')).strip('\\r\\n'),6:bOrNot(packet.sprintf('%SNMP.community%')),7:bOrNot(packet.sprintf('%Padding.load%'))}
    return dic

def NullOrNot(packet):
    return packet.sprintf('%Raw.load%')[1:-1]+packet.sprintf('%SMBNegociate_Protocol_Request_Tail.BufferData%')[1:-1]+packet.sprintf('%Padding.Load%')[1:-1]+packet.sprintf('%SNMP.community%')[1:-1]

def HttpFilter(packet,Num,postkey='PostData'):
    if re.match(r'\w+ ',packet.sprintf('%Raw.load%')[1:-1]):#匹配到，是请求头
        onestr='No.:'+str(Num)+'\\r\\n'+'sport:'+packet.sprintf('%TCP.sport%')+'\\r\\n'+packet.sprintf('%Raw.load%')[1:-1]
        onelist=onestr.strip('\\r\\n').replace('\\r\\n\\r\\n','\\r\\n{}:'.format(postkey),1).split('\\r\\n')
        onelist[2]=onelist[2].replace(' ',':',1)
    elif packet.sprintf('%Raw.load%')[1:-1]!='':#未匹配到，不等于空，是PostData
        onestr='No.:'+str(Num)+'\\r\\n'+'sport:'+packet.sprintf('%TCP.sport%')+'\\r\\n'+postkey+':'+packet.sprintf('%Raw.load%')[1:-1]
        onelist=onestr.strip('\\r\\n').split('\\r\\n')
    else:onelist=''#数据为空
    return onelist

def ListToDic(onelist):
    dic,s={},''
    while '' in onelist:onelist.remove('')
    for i in range(len(onelist)):
        if 'PostData:' in onelist[i]:
            onelist[i]='\n'.join(onelist[i:])
            dic[HttpFormHeader.negative[onelist[i].split(':',1)[0].lower()]]=onelist[i].split(':',1)[1]
            break
        elif onelist[i].split(':',1)[0].lower() not in HttpFormHeader.negative.keys():
            s=s+onelist[i]+';'
            dic[HttpFormHeader.negative['otherkeys']]=s
        else:dic[HttpFormHeader.negative[onelist[i].split(':',1)[0].lower()]]=onelist[i].split(':',1)[1]
    return dic

def userStyles(*sheets):
    for sheet in sheets:
        for irow, row in enumerate(sheet.rows, start=1):
            count=0
            for cell in row:
                count+=1
                if count>12:count=1
                if irow==1:
                    cell.font=Font(name='Times New Roman',bold=True)
                    cell.fill=PatternFill(patternType='solid',fgColor='FFFF00')
                else:cell.font = Font(name='Times New Roman')
                ProgressBar(count,startTime)

def ProgressBar(i,startTime,word='处理数据'):
    words,m,timer='正在{}'.format(word),'.'*i,time.clock()-startTime
    sys.stdout.write(words+m+' '*(13-i)+time.strftime('%M:%S',time.gmtime(timer))+'\r')
    sys.stdout.flush()

def bOrNot(strings):
    if strings.startswith('b'):string=strings
    else:string=strings[1:-1]
    return string

class FormHeader(object):
    def __init__(self,FormHeader):
        self.positive=dict(zip(range(1,len(FormHeader)+1),FormHeader))
        self.negative=dict(zip(list(map(str.lower,FormHeader)),range(1,len(FormHeader)+1)))

HttpFormHeader=FormHeader(['No.','Sport','GET','POST','PostData','Host','Accept','Accept-Charset','Accept-Encoding','Accept-Language',
'Accept-Ranges','Authorization','Cache-Control','Connection','Cookie','Content-Length','Content-Type','Date','Expect','From',
'Max-Forwards','Pragma','Proxy-Authorization','Range','Referer','Upgrade','User-Agent','Warning','OtherKeys'])
PublicFormHeader=FormHeader(['No.','Sport','Dport','Protocol','RawLoad','BufferData','PaddingLoad'])
UDPFormHeader=FormHeader(['No.','Sport','Dport','Protocol','RawLoad','SNMP_Community','PaddingLoad'])
path,filename=os.getcwd()+'\\',input('输入pcap文件名（脚本所在文件夹）,空格分割：')
#print(filename)
startTime=time.clock()
for i in filename.split(' '):main(path+i)